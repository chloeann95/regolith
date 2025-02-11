import os
import sys

import habanero
import openpyxl
import pytest
from xonsh.lib import subprocess

from regolith.broker import load_db
from regolith.main import main

builder_map = [
    "cv",
    "review-prop",
    "annual-activity",
    "beamplan",
    "current-pending",
    "figure",
    "html",
    "internalhtml",
    "preslist",
    "publist",
    "recent-collabs",
    "beamplan",
    "grantreport",
    "resume",
    "review-man",
    "reimb"
]
db_srcs = ["mongo", "fs"]

xls_check = ("B17", "B20", "B36")
recent_collabs_xlsx_check = ["A51", "B51", "C51"]


def is_same(text0: str, text1: str, ignored: list):
    """Compare the content of two text. If there are different words in text0 and text1 and the word in text0
    does not contain ignored substring, return False. Else return True."""

    def should_ignore(word):
        for w in ignored:
            if w in word:
                return True
        return False

    words0, words1 = text0.split(), text1.split()
    if len(words0) != len(words1):
        return False
    for word0, word1 in zip(words0, words1):
        if not should_ignore(word0) and word0 != word1:
            return False
    return True


def prep_figure():
    # Make latex file with some jinja2 in it
    text = r"""
    \include{ {{-get_file_path(db['groups']['ergs'], 'hello')-}}}"""
    with open("figure.tex", "w") as f:
        f.write(text)
    # make file to be loaded
    os.makedirs("fig", exist_ok=True)
    with open("fig/hello.txt", "w") as f:
        f.write("hello world")
    # load the db and register the file
    db = load_db()
    print(db.get_file_path(db["groups"]["ergs"], "hello"))
    if not db.get_file_path(db["groups"]["ergs"], "hello"):
        db.add_file(db["groups"]["ergs"], "hello", "fig/hello.txt")


@pytest.mark.skipif(sys.platform == "win32", reason="does not run on windows")
@pytest.mark.parametrize("bm", builder_map)
@pytest.mark.parametrize("db_src", db_srcs)
def test_builder(bm, db_src, make_db, make_mongodb, monkeypatch):
    # FIXME: Somehow the mongo backend failed to build figure
    # FIXME: now fs is failing to build figure
    # if db_src == "mongo" and bm == "figure":
    if bm == "figure":
        return
    if db_src == "fs":
        repo = make_db
    elif db_src == "mongo":
        if make_mongodb is False:
            pytest.skip("Mongoclient failed to start")
        else:
            repo = make_mongodb
    else:
        raise ValueError("Unknown database source: {}".format(db_src))
    os.chdir(repo)
    if bm == "figure":
        prep_figure()
    if bm == "internalhtml":
        # for some reason the mocking of the crossref call doesn't work when the
        # test is run using subprocess, so skip in this case.
        # the functionality is fully tested in test_builder_python
        pytest.skip("mocking of Crossref not working with subprocess")
    if bm == "html":
        os.makedirs("templates/static", exist_ok=True)
    if bm == "reimb" or bm == "recent-collabs":
        subprocess.run(["regolith", "build", bm, "--no-pdf", "--people",
                        "scopatz"], check=True, cwd=repo)
    elif bm == "annual-activity":
        subprocess.run(["regolith", "build", bm, "--no-pdf", "--people",
                        "sbillinge", "--from", "2017-04-01"], check=True, cwd=repo)
    elif bm == "grantreport":
        main(["build", bm, "--no-pdf", "--grant", "SymPy-1.1",
              "--from", "2017-04-01", "--to", "2018-03-31"])
    else:
        subprocess.run(["regolith", "build", bm, "--no-pdf"], check=True, cwd=repo)
    os.chdir(os.path.join(repo, "_build", bm))
    expected_base = os.path.join(os.path.dirname(__file__), "outputs")
    for root, dirs, files in os.walk("."):
        for file in files:
            if file in os.listdir(os.path.join(expected_base, bm, root)):
                fn1 = os.path.join(repo, "_build", bm, root, file)
                if bm == "reimb":
                    actual = openpyxl.load_workbook(fn1)["T&B"]
                    actual = [str(actual[b]) for b in xls_check]
                elif bm == "recent-collabs":
                    if 'nsf' in fn1:
                        sheet = "NSF COA Template"
                    else:
                        sheet = "Collaborators"
                    actual = openpyxl.load_workbook(fn1)[sheet]
                    actual = [str(actual[cell]) for cell in recent_collabs_xlsx_check]
                else:
                    with open(fn1, "r") as f:
                        actual = f.read()
                fn2 = os.path.join(expected_base, bm, root, file)
                if bm == "reimb":
                    expected = openpyxl.load_workbook(fn2)["T&B"]
                    expected = [str(expected[b]) for b in xls_check]
                elif bm == "recent-collabs":
                    if 'nsf' in fn2:
                        sheet = "NSF COA Template"
                    else:
                        sheet = "Collaborators"
                    expected = openpyxl.load_workbook(fn2)[sheet]
                    expected = [str(expected[cell]) for cell in recent_collabs_xlsx_check]
                else:
                    with open(fn2, "r") as f:
                        expected = f.read()

                # Skip because of a date time in
                if file != "rss.xml":
                    if file.endswith('.html') or file.endswith('.tex'):
                        if not is_same(expected, actual, ['../..', 'tmp']):
                            assert actual == expected
                    else:
                        assert actual == expected


@pytest.mark.parametrize("db_src", db_srcs)
@pytest.mark.parametrize("bm", builder_map)
def test_builder_python(bm, db_src, make_db, make_mongodb,
                        monkeypatch):
    # FIXME: Somehow the mongo backend failed to build figure
    # FIXME: now fs is failing to build figure
    # if db_src == "mongo" and bm == "figure":
    if bm == "figure":
        return
    if db_src == "fs":
        repo = make_db
    elif db_src == "mongo":
        if make_mongodb is False:
            pytest.skip("Mongoclient failed to start")
        else:
            repo = make_mongodb
    os.chdir(repo)
    if bm == "figure":
        prep_figure()
    if bm == "internalhtml":
        def mockreturn(*args, **kwargs):
            mock_article = {'message': {'author': [{"given":"SJL","family":"B"}],
                                        "short-container-title": ["J Club Paper"],
                                        "volume": 10,
                                        "title": ["title"],
                                        "issued": {"date-parts":[[1971]]}}
                            }
            return mock_article
        monkeypatch.setattr(habanero.Crossref, "works", mockreturn)
    if bm == "html":
        os.makedirs("templates/static", exist_ok=True)
    if bm == "reimb" or bm == "recent-collabs":
        main(["build", bm, "--no-pdf", "--people", "scopatz"])
    elif bm == "annual-activity":
        main(["build", bm, "--no-pdf", "--people",
              "sbillinge", "--from", "2017-04-01"])
    elif bm == "grantreport":
        main(["build", bm, "--no-pdf", "--grant", "SymPy-1.1",
              "--to", "2018-03-31"])
    else:
        main(["build", bm, "--no-pdf"])
    os.chdir(os.path.join(repo, "_build", bm))
    expected_base = os.path.join(os.path.dirname(__file__), "outputs")
    for root, dirs, files in os.walk("."):
        for file in files:
            if file in os.listdir(os.path.join(expected_base, bm, root)):
                html_tex_bool = False
                if file.endswith('.html') or file.endswith('.tex'):
                    html_tex_bool = True
                fn1 = os.path.join(repo, "_build", bm, root, file)
                if bm == "reimb":
                    actual = openpyxl.load_workbook(fn1)["T&B"]
                    actual = [str(actual[b]) for b in xls_check]
                elif bm == "recent-collabs":
                    if 'nsf' in fn1:
                        sheet = "NSF COA Template"
                    else:
                        sheet = "Collaborators"
                    actual = openpyxl.load_workbook(fn1)[sheet]
                    actual = [str(actual[cell]) for cell in recent_collabs_xlsx_check]
                elif html_tex_bool:
                    with open(fn1, "r") as f:
                        actual = [line.strip() for line in f]
                    actual = [string for string in actual if '..' not in string]
                    actual = [string for string in actual if
                              'Temp' not in string]
                    actual = [string for string in actual if
                              '/tmp/' not in string]
                    actual = [string for string in actual if
                                len(string) != 0]
                else:
                    with open(fn1, "r") as f:
                        actual = f.read()
                fn2 = os.path.join(expected_base, bm, root, file)
                if bm == "reimb":
                    expected = openpyxl.load_workbook(fn2)["T&B"]
                    expected = [str(expected[b]) for b in xls_check]
                elif bm == "recent-collabs":
                    if 'nsf' in fn2:
                        sheet = "NSF COA Template"
                    else:
                        sheet = "Collaborators"
                    expected = openpyxl.load_workbook(fn2)[sheet]
                    expected = [str(expected[cell]) for cell in recent_collabs_xlsx_check]
                elif html_tex_bool:
                    with open(fn2, "r") as f:
                        expected = [line.strip() for line in f]
                    expected = [string for string in expected if
                              '..' not in string]
                    expected = [string for string in expected if
                              'Temp' not in string]
                    expected = [string for string in expected if
                              '/tmp/' not in string]
                    expected = [string for string in expected if
                                len(string) != 0]
                else:
                    with open(fn2, "r") as f:
                        expected = f.read()

                # Skip because of a date time in
                if file != "rss.xml":
                    assert expected == actual
